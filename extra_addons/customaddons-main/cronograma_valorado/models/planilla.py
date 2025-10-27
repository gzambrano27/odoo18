import base64         # Para codificar y decodificar datos en base64
import io             # Para trabajar con flujos de datos en memoria (buffers)
import xlsxwriter     # Para generar archivos Excel
import matplotlib     # Para generar gráficos
import matplotlib.pyplot as plt  # Interfaz de pyplot para graficar
matplotlib.use('Agg')  # Configura un backend no interactivo (útil en servidores sin entorno gráfico)
from openpyxl import load_workbook  # Para leer archivos Excel
from openpyxl import Workbook       # Para crear archivos Excel (no se usa en este código, pero se importa)
from odoo import models, fields, api, _  # Importa modelos, campos, API y traducción en Odoo
import logging        # Para el registro (logs)
_logger = logging.getLogger(__name__)  # Configura el logger para este módulo
from openpyxl.utils.datetime import from_excel  # Para convertir fechas de Excel a datetime
from odoo.exceptions import UserError  # Para lanzar excepciones de usuario en Odoo
import numpy as np    # Para trabajar con arrays numéricos
from scipy.interpolate import make_interp_spline  # Para interpolar suavemente puntos de datos (curva S)
from datetime import datetime, timedelta  # Para trabajar con fechas y tiempos

class Planilla(models.Model):
    _name = 'cronograma.planilla'  # Nombre técnico del modelo
    _inherit = ['portal.mixin', 'mail.thread', 'mail.activity.mixin']  # Hereda funcionalidades para portal, chatter y actividades
    _description = 'Planilla'  # Descripción del modelo

    # Definición de campos básicos
    name = fields.Char(string='Nombre', copy=False, default='PLANILLA - New')  # Nombre de la planilla
    usuario_id = fields.Many2one(
        'res.users',
        string='Usuario',
        tracking=True,
        default=lambda self: self.env.user,
        readonly=True
    )  # Usuario asociado, se establece por defecto el usuario actual
    cuenta_analitica_id = fields.Many2one('account.analytic.account', string='Cuenta Analítica', tracking=True)  # Cuenta analítica asociada
    notebook_ids = fields.One2many('cronograma.notebook', 'planilla_id', string='Notebook', tracking=True)  # Relación uno a muchos con los notebooks

    # Campos para totales (calculados a partir de los notebooks)
    total_subtotal = fields.Monetary(string='Total Valorado', compute='_compute_totales', store=True, readonly=True)
    total_diferencia_subtotal = fields.Monetary(string='Total Diferencia', compute='_compute_totales', store=True, readonly=True)
    total_subtotal_planilla = fields.Monetary(string='Total Avance Real', compute='_compute_totales', store=True, readonly=True)

    # Información de la compañía y moneda
    company_id = fields.Many2one("res.company", string="Compañía", copy=False,
                                 default=lambda self: self.env.company, _check_company_auto=True)
    currency_id = fields.Many2one(related="company_id.currency_id", string="Moneda", store=False, readonly=True)

    @api.depends('notebook_ids.subtotal', 'notebook_ids.subtotal_planilla')
    def _compute_totales(self):
        """Calcula los totales sumando los subtotales y avances de cada notebook asociado."""
        for record in self:
            record.total_subtotal = sum(notebook.subtotal for notebook in record.notebook_ids)
            record.total_diferencia_subtotal = sum(notebook.diferencia_subtotal for notebook in record.notebook_ids)
            record.total_subtotal_planilla = sum(notebook.subtotal_planilla for notebook in record.notebook_ids)

    # Campo de estado con sus posibles valores
    state = fields.Selection(
        [('draft', 'Borrador'),
         ('to_approve', 'Por Aprobar'),
         ('approved', 'Aprobado'),
         ('cancel', 'Cancelado')],
        string='Estado', default='draft', tracking=True, readonly=True, copy=False
    )
    

    @api.model
    def create(self, vals):
        """Sobrescribe el método create para asignar secuencialmente el nombre y definir estado y usuario."""
        if vals.get('name', 'PLANILLA - New') == 'PLANILLA - New':
            vals['name'] = self.env['ir.sequence'].next_by_code('cronograma.planilla') or 'PLANILLA - New'
        vals['state'] = 'draft'
        vals['usuario_id'] = self.env.user.id
        return super(Planilla, self).create(vals)

    # Métodos para cambiar el estado de la planilla
    def action_to_approve(self):
        """Cambia el estado a 'Por Aprobar'."""
        self.write({'state': 'to_approve'})

    def action_approve(self):
        """Cambia el estado a 'Aprobado'."""
        self.write({'state': 'approved'})

    def action_cancel(self):
        """Cambia el estado a 'Cancelado'."""
        self.write({'state': 'cancel'})

    def action_reset_to_draft(self):
        """Reestablece el estado a 'Borrador'."""
        self.write({'state': 'draft'})

    def export_to_excel(self):
        """Genera y exporta un archivo Excel detallado con los datos de la planilla."""
        for planilla in self:
            # Verifica que la planilla tenga al menos un notebook asociado
            if not planilla.notebook_ids:
                raise UserError("La planilla no tiene Notebooks para exportar.")

            # Crea un buffer en memoria para guardar el archivo Excel
            buffer = io.BytesIO()
            # Crea un workbook Excel en memoria
            workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
            # Agrega una hoja al workbook
            worksheet = workbook.add_worksheet('Planilla')

            # Define formatos para el Excel
            header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1})
            text_format = workbook.add_format({'border': 1})
            number_format = workbook.add_format({'border': 1, 'num_format': '#,##0.00'})
            date_format = workbook.add_format({'border': 1, 'num_format': 'dd/mm/yyyy'})

            # Encabezados principales para el Excel
            headers = [
                'Planilla Nombre',
                'Rubro',
                'Descripción',
                'Precio Unitario',
                'Cantidad',
                'Subtotal',
                'Consumo',
                'Diferencia Cantidad',
                'Subtotal Planilla',
                'Diferencia Subtotal',
                'Observación'
            ]

            # Se añaden encabezados dinámicos para los detalles (cantidad variable de columnas)
            max_detalles = max(len(notebook.detalle_ids) for notebook in planilla.notebook_ids)
            for i in range(1, max_detalles + 1):
                headers.extend([
                    f'Detalle {i} - Fecha',
                    f'Detalle {i} - Cantidad',
                    f'Detalle {i} - Valor',
                    f'Detalle {i} - Cantidad Avance Real',
                    f'Detalle {i} - Valor Avance Real'
                ])

            # Escribe los encabezados en la primera fila
            for col_num, header in enumerate(headers):
                worksheet.write(0, col_num, header, header_format)

            # Recorre cada notebook y escribe sus datos
            row = 1  # Inicia en la fila 2 (índice 1)
            for notebook in planilla.notebook_ids:
                # Nota: ahora se utiliza el campo 'rubro' (Char) y no 'rubro'
                notebook_data = [
                    planilla.name,  # Nombre de la planilla
                    notebook.rubro or '',  # Rubro como texto
                    notebook.descripcion or '',  # Descripción del notebook
                    notebook.precio_unitario or 0.0,  # Precio unitario
                    notebook.cantidad or 0.0,  # Cantidad total
                    notebook.subtotal or 0.0,  # Subtotal calculado
                    notebook.consumo or 0.0,  # Consumo acumulado
                    notebook.diferencia_cantidad or 0.0,  # Diferencia de cantidad
                    notebook.subtotal_planilla or 0.0,  # Subtotal de avance real
                    notebook.diferencia_subtotal or 0.0,  # Diferencia de subtotal
                    notebook.observacion or ''  # Observaciones
                ]

                # Procesa los detalles asociados al notebook
                detalles = []
                for detalle in notebook.detalle_ids:
                    # Formatea la fecha a 'dd/mm/yyyy'
                    fecha_formateada = detalle.fecha.strftime('%d/%m/%Y') if detalle.fecha else ''
                    detalles.extend([
                        fecha_formateada,
                        detalle.cantidad or 0.0,
                        detalle.valor or 0.0,
                        detalle.cantidad_avance_real or 0.0,
                        detalle.valor_avance_real or 0.0
                    ])

                # Si un notebook tiene menos detalles que el máximo, se rellenan celdas vacías
                while len(detalles) < max_detalles * 5:
                    detalles.extend(['', '', '', '', ''])

                # Combina los datos básicos con los detalles
                row_data = notebook_data + detalles

                # Escribe cada celda de la fila en el Excel
                for col_num, cell_data in enumerate(row_data):
                    if isinstance(cell_data, (int, float)):
                        worksheet.write(row, col_num, cell_data, number_format)
                    # Se identifica si la cadena parece una fecha (contiene '/')
                    elif isinstance(cell_data, str) and '/' in cell_data:
                        worksheet.write(row, col_num, cell_data, date_format)
                    else:
                        worksheet.write(row, col_num, cell_data, text_format)
                row += 1

            # Finaliza y cierra el workbook, y vuelve al inicio del buffer
            workbook.close()
            buffer.seek(0)

            # Lee el contenido del buffer para crear un attachment
            file_data = buffer.read()
            buffer.close()

            # Crea un attachment en Odoo para el archivo generado
            attachment = self.env['ir.attachment'].create({
                'name': f'{planilla.name}.xlsx',
                'type': 'binary',
                'datas': base64.b64encode(file_data).decode('utf-8'),
                'res_model': self._name,
                'res_id': planilla.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            # Retorna una acción que permite descargar el archivo
            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'new',
            }

    def export_to_excel_project(self):
        """Exporta un Excel con columnas específicas para proyectos."""
        for planilla in self:
            if not planilla.notebook_ids:
                raise UserError("La planilla no tiene Notebooks para exportar.")

            buffer = io.BytesIO()
            workbook = xlsxwriter.Workbook(buffer, {'in_memory': True})
            worksheet = workbook.add_worksheet('Planilla')

            header_format = workbook.add_format({'bold': True, 'bg_color': '#D3D3D3', 'border': 1})
            text_format = workbook.add_format({'border': 1})
            number_format = workbook.add_format({'border': 1, 'num_format': '#,##0.00'})
            date_format = workbook.add_format({'border': 1, 'num_format': 'dd/mm/yyyy'})

            # Encabezados específicos para este reporte
            headers = [
                'Rubro',
                'Fecha de Inicio',
                'Fecha de Finalización',
                'Precio Unitario',
                'Cantidad Total'
            ]

            for col_num, header in enumerate(headers):
                worksheet.write(0, col_num, header, header_format)

            row = 1
            for notebook in planilla.notebook_ids:
                # Se utiliza el campo 'rubro' (Char)
                rubro = notebook.rubro or ''
                # Se recopilan las fechas de los detalles
                fechas = [detalle.fecha for detalle in notebook.detalle_ids if detalle.fecha]
                if fechas:
                    fecha_inicio = min(fechas).strftime('%d/%m/%Y')
                    fecha_fin = max(fechas).strftime('%d/%m/%Y')
                else:
                    fecha_inicio = ''
                    fecha_fin = ''

                precio_unitario = notebook.precio_unitario or 0.0
                cantidad_total = notebook.cantidad or 0.0

                row_data = [
                    rubro,
                    fecha_inicio,
                    fecha_fin,
                    precio_unitario,
                    cantidad_total
                ]

                for col_num, cell_data in enumerate(row_data):
                    if col_num in [3, 4]:
                        worksheet.write(row, col_num, cell_data, number_format)
                    elif col_num in [1, 2] and cell_data:
                        worksheet.write(row, col_num, cell_data, date_format)
                    else:
                        worksheet.write(row, col_num, cell_data, text_format)
                row += 1

            workbook.close()
            buffer.seek(0)

            file_data = buffer.read()
            buffer.close()

            attachment = self.env['ir.attachment'].create({
                'name': f'{planilla.name}.xlsx',
                'type': 'binary',
                'datas': base64.b64encode(file_data),
                'res_model': self._name,
                'res_id': planilla.id,
                'mimetype': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'new',
            }

    # Campos para importar archivos Excel
    import_file = fields.Binary(string="Archivo Excel")
    import_filename = fields.Char(string="Nombre del Archivo")

    def import_from_excel(self):
        """Importa datos desde un archivo Excel al modelo cronograma.planilla."""
        # Verifica que se haya cargado un archivo y que tenga extensión .xlsx
        if not self.import_file or not self.import_filename.endswith('.xlsx'):
            raise UserError(_("Por favor, cargue un archivo Excel válido."))

        try:
            # Decodifica el archivo cargado y crea un buffer
            file_data = base64.b64decode(self.import_file)
            file_content = io.BytesIO(file_data)
            # Carga el workbook con openpyxl
            workbook = load_workbook(file_content)
            sheet = workbook.active  # Selecciona la primera hoja
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise UserError(_("El archivo Excel está vacío."))

            headers = rows[0]  # La primera fila son los encabezados
            data_rows = rows[1:]  # El resto son datos
            planilla = self  # Trabaja sobre la planilla actual

            for row in data_rows:
                if len(row) < len(headers):
                    raise UserError(_("El archivo Excel tiene filas con menos columnas que los encabezados."))

                try:
                    # Si no hay valor en la primera columna, se omite la fila
                    if not row[0]:
                        continue

                    # Crea un registro de notebook utilizando los datos del Excel
                    notebook_data = {
                        'planilla_id': planilla.id,
                        'rubro': row[1],  # Asigna directamente el valor al campo rubro (Char)
                        'descripcion': row[2],
                        'precio_unitario': float(row[3]) if row[3] else 0.0,
                        'cantidad': float(row[4]) if row[4] else 0.0,
                        'subtotal': float(row[5]) if row[5] else 0.0,
                        'consumo': float(row[6]) if row[6] else 0.0,
                        'diferencia_cantidad': float(row[7]) if row[7] else 0.0,
                        'subtotal_planilla': float(row[8]) if row[8] else 0.0,
                        'diferencia_subtotal': float(row[9]) if row[9] else 0.0,
                        'observacion': row[10],
                    }
                    notebook = planilla.notebook_ids.create(notebook_data)

                    # Procesa los detalles (cada detalle ocupa 5 columnas)
                    detalle_index = 11
                    while detalle_index < len(headers):
                        if "Detalle" not in headers[detalle_index]:
                            break
                        try:
                            # Procesa y formatea la fecha del detalle
                            detalle_fecha = row[detalle_index]
                            if isinstance(detalle_fecha, datetime):
                                detalle_fecha = detalle_fecha.strftime('%Y-%m-%d')
                            elif isinstance(detalle_fecha, (int, float)):
                                detalle_fecha = from_excel(detalle_fecha).strftime('%Y-%m-%d')
                            elif isinstance(detalle_fecha, str) and detalle_fecha.strip():
                                detalle_fecha = datetime.strptime(detalle_fecha, '%d/%m/%Y').strftime('%Y-%m-%d')
                            else:
                                detalle_fecha = None

                            if not detalle_fecha:
                                raise ValueError("La fecha es inválida o está vacía.")

                            detalle_data = {
                                'notebook_id': notebook.id,
                                'fecha': detalle_fecha,
                                'cantidad': float(row[detalle_index + 1]) if row[detalle_index + 1] else 0.0,
                                'valor': float(row[detalle_index + 2]) if row[detalle_index + 2] else 0.0,
                                'cantidad_avance_real': float(row[detalle_index + 3]) if detalle_index + 3 < len(row) and row[detalle_index + 3] else 0.0,
                                'valor_avance_real': float(row[detalle_index + 4]) if detalle_index + 4 < len(row) and row[detalle_index + 4] else 0.0,
                            }
                            notebook.detalle_ids.create(detalle_data)
                        except Exception as e:
                            detalle_index += 5  # Si ocurre error, salta al siguiente grupo de 5 columnas
                            continue
                        detalle_index += 5

                except Exception as e:
                    continue

            file_content.close()
            return True

        except Exception as e:
            raise UserError(_("Error al procesar el archivo Excel: %s") % str(e))

    def import_from_excel_project(self):
        """Importa datos desde un archivo Excel para proyectos e interpreta niveles."""
        if not self.import_file or not self.import_filename.endswith('.xlsx'):
            raise UserError("Por favor, cargue un archivo Excel válido.")

        def parse_date(fecha):
            fecha_str = str(fecha).strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(fecha_str, fmt)
                except Exception:
                    continue
            raise UserError("La fecha '%s' no tiene un formato válido." % fecha_str)

        try:
            file_data = base64.b64decode(self.import_file)
            file_content = io.BytesIO(file_data)
            workbook = load_workbook(file_content)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise UserError("El archivo Excel está vacío.")

            headers = rows[0]
            header_index = {header: idx for idx, header in enumerate(headers)}
            required_columns = [
                "Estructura de desglose de trabajo",
                "Nombre",
                "Costo",
                "Cantidad",
                "Inicio",
                "Finalizar"
            ]
            for col in required_columns:
                if col not in header_index:
                    raise UserError("El archivo Excel no contiene la columna '%s'." % col)

            data_rows = rows[1:]
            planilla = self

            for row in data_rows:
                if len(row) < len(headers):
                    raise UserError("El archivo Excel tiene filas con menos columnas que los encabezados.")

                estructura = row[header_index["Estructura de desglose de trabajo"]]
                if not estructura:
                    continue
                estructura_str = str(estructura).strip()
                nivel = estructura_str.count('.') + 1

                descripcion = row[header_index["Nombre"]]
                rubro = descripcion

                precio_unitario_val = row[header_index["Costo"]]
                precio_unitario = float(precio_unitario_val) if precio_unitario_val is not None else 0.0

                cantidad_val = row[header_index["Cantidad"]]
                cantidad_total = float(cantidad_val) if cantidad_val is not None else 0.0

                fecha_inicio = parse_date(row[header_index["Inicio"]])
                fecha_final = parse_date(row[header_index["Finalizar"]])

                notebook_data = {
                    'planilla_id': planilla.id,
                    'rubro': rubro,
                    'descripcion': descripcion,
                    'precio_unitario': precio_unitario,
                    'cantidad': cantidad_total,
                    'nivel': nivel
                }
                notebook = planilla.notebook_ids.create(notebook_data)

                if nivel < 4:
                    continue

                total_dias = (fecha_final - fecha_inicio).days + 1
                if total_dias <= 0:
                    raise UserError("El rango de fechas no es válido.")

                cantidad_diaria = round(cantidad_total / total_dias, 2)
                total_distribuido = round(cantidad_diaria * total_dias, 2)
                sobrante = round(cantidad_total - total_distribuido, 2)

                current_date = fecha_inicio
                for dia in range(total_dias):
                    detalle_data = {
                        'notebook_id': notebook.id,
                        'fecha': current_date.strftime('%Y-%m-%d'),
                        'cantidad': cantidad_diaria,
                        'valor': round(cantidad_diaria * precio_unitario, 2),
                        'cantidad_avance_real': 0.0,
                        'valor_avance_real': 0.0,
                    }
                    if dia == total_dias - 1:
                        detalle_data['cantidad'] += sobrante
                        detalle_data['valor'] = round(detalle_data['cantidad'] * precio_unitario, 2)

                    notebook.detalle_ids.create(detalle_data)
                    current_date += timedelta(days=1)

            file_content.close()
            return True

        except Exception as e:
            raise UserError("Error al procesar el archivo Excel: %s" % str(e))
        

    def import_from_excel_projectXXNEW(self):
        """Importa datos desde un archivo Excel para proyectos al modelo cronograma.planilla.
           Se importan únicamente las tareas hijos, identificadas por la columna
           'Estructura de desglose de trabajo' que debe tener al menos dos puntos.
           Además, todos los datos se obtienen usando el nombre de la columna."""
        if not self.import_file or not self.import_filename.endswith('.xlsx'):
            raise UserError(_("Por favor, cargue un archivo Excel válido."))

        def parse_date(fecha):
            """Intenta convertir una fecha a datetime usando dos formatos: '%d/%m/%Y' y '%Y-%m-%d'."""
            fecha_str = str(fecha).strip()
            for fmt in ("%d/%m/%Y", "%Y-%m-%d"):
                try:
                    return datetime.strptime(fecha_str, fmt)
                except Exception:
                    continue
            raise UserError(_("La fecha '%s' no tiene un formato válido.") % fecha_str)

        try:
            file_data = base64.b64decode(self.import_file)
            file_content = io.BytesIO(file_data)
            workbook = load_workbook(file_content)
            sheet = workbook.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise UserError(_("El archivo Excel está vacío."))

            # Obtener la fila de encabezados y crear un diccionario para acceder por nombre
            headers = rows[0]
            header_index = {header: idx for idx, header in enumerate(headers)}

            # Verificar que existan las columnas requeridas
            required_columns = [
                "Estructura de desglose de trabajo",
                "Nombre",
                "Costo",
                "Cantidad",
                "Inicio",
                "Finalizar"
            ]
            for col in required_columns:
                if col not in header_index:
                    raise UserError(_("El archivo Excel no contiene la columna '%s'.") % col)

            data_rows = rows[1:]
            planilla = self

            for row in data_rows:
                if len(row) < len(headers):
                    raise UserError(_("El archivo Excel tiene filas con menos columnas que los encabezados."))

                # Filtrar solo las tareas hijos basándose en la columna "Estructura de desglose de trabajo"
                estructura = row[header_index["Estructura de desglose de trabajo"]]
                if not estructura:
                    continue
                estructura_str = str(estructura).strip()
                # Se considera tarea hijo si tiene al menos dos puntos (por ejemplo, "1.1.1")
                if estructura_str.count('.') < 2:
                    continue

                # Obtener los datos usando el nombre de columna
                descripcion = row[header_index["Nombre"]]
                rubro = descripcion

                precio_unitario_val = row[header_index["Costo"]]
                precio_unitario = float(precio_unitario_val) if precio_unitario_val is not None else 0.0

                cantidad_val = row[header_index["Cantidad"]]
                cantidad_total = float(cantidad_val) if cantidad_val is not None else 0.0

                fecha_inicio = row[header_index["Inicio"]]
                fecha_final = row[header_index["Finalizar"]]

                # Convertir las fechas a datetime usando la función parse_date
                fecha_inicio = parse_date(fecha_inicio)
                fecha_final = parse_date(fecha_final)

                # Crear el notebook con los datos del registro
                notebook_data = {
                    'planilla_id': planilla.id,
                    'rubro': rubro,
                    'descripcion': descripcion,
                    'precio_unitario': precio_unitario,
                    'cantidad': cantidad_total,
                }
                notebook = planilla.notebook_ids.create(notebook_data)
                _logger.info("Notebook creado con ID: %s", notebook.id)

                total_dias = (fecha_final - fecha_inicio).days + 1
                if total_dias <= 0:
                    raise UserError(_("El rango de fechas no es válido."))

                cantidad_diaria = round(cantidad_total / total_dias, 2)
                total_distribuido = round(cantidad_diaria * total_dias, 2)
                sobrante = round(cantidad_total - total_distribuido, 2)

                current_date = fecha_inicio
                # Crear un detalle para cada día del rango de fechas
                for dia in range(total_dias):
                    detalle_data = {
                        'notebook_id': notebook.id,
                        'fecha': current_date.strftime('%Y-%m-%d'),
                        'cantidad': cantidad_diaria,
                        'valor': round(cantidad_diaria * precio_unitario, 2),
                        'cantidad_avance_real': 0.0,
                        'valor_avance_real': 0.0,
                    }
                    # En el último día, ajustar la cantidad con el sobrante
                    if dia == total_dias - 1:
                        detalle_data['cantidad'] += sobrante
                        detalle_data['valor'] = round(detalle_data['cantidad'] * precio_unitario, 2)

                    notebook.detalle_ids.create(detalle_data)
                    _logger.info("Detalle creado para la fecha: %s", current_date.strftime('%Y-%m-%d'))
                    current_date += timedelta(days=1)

            file_content.close()
            _logger.info("Importación completada con éxito.")
            return True

        except Exception as e:
            _logger.error("Error al procesar el archivo Excel: %s", str(e))
            raise UserError(_("Error al procesar el archivo Excel: %s") % str(e))

    def generate_curva_s(self):
        """Genera un gráfico de la curva S comparativa basado en el subtotal acumulado y el avance real acumulado.
        Solo se anotan los valores cuando éstos cambian respecto al punto anterior.
        Cada anotación se muestra en un recuadro del color de la línea correspondiente."""
        for planilla in self:
            if not planilla.notebook_ids:
                raise UserError("No hay datos suficientes para graficar la curva S.")

            # Diccionarios para acumular subtotales y avances reales por fecha
            subtotales_por_fecha = {}
            avance_real_por_fecha = {}

            # Recorre cada notebook y cada detalle para acumular los valores por fecha
            for notebook in planilla.notebook_ids:
                for detalle in notebook.detalle_ids:
                    if detalle.fecha:
                        subtotal = detalle.cantidad * notebook.precio_unitario
                        avance_real = detalle.cantidad_avance_real * notebook.precio_unitario
                        if detalle.fecha not in subtotales_por_fecha:
                            subtotales_por_fecha[detalle.fecha] = 0
                            avance_real_por_fecha[detalle.fecha] = 0
                        subtotales_por_fecha[detalle.fecha] += subtotal
                        avance_real_por_fecha[detalle.fecha] += avance_real

            total_acumulado = planilla.total_subtotal

            # Ordena las fechas y calcula los porcentajes acumulados para cada fecha
            fechas = sorted(subtotales_por_fecha.keys())
            valores_acumulados = []
            avance_real_acumulado = []
            acumulado = 0
            real_acumulado = 0

            for fecha in fechas:
                acumulado += subtotales_por_fecha[fecha]
                real_acumulado += avance_real_por_fecha[fecha]
                porcentaje_acumulado = round((acumulado / total_acumulado) * 100, 2) if total_acumulado else 0
                porcentaje_real_acumulado = round((real_acumulado / total_acumulado) * 100, 2) if total_acumulado else 0
                valores_acumulados.append(porcentaje_acumulado)
                avance_real_acumulado.append(porcentaje_real_acumulado)

            if not fechas or not valores_acumulados or not avance_real_acumulado:
                raise UserError("No se pudo generar la curva S debido a datos insuficientes.")

            # Crea la figura y el eje para el gráfico
            fig, ax = plt.subplots(figsize=(22, 11))

            # Convierte las fechas a índices numéricos para la interpolación
            x_vals = np.arange(len(fechas))
            spline_acumulado = make_interp_spline(x_vals, np.array(valores_acumulados), k=3)
            spline_avance_real = make_interp_spline(x_vals, np.array(avance_real_acumulado), k=3)

            # Genera puntos interpolados para suavizar las curvas
            x_smooth = np.linspace(x_vals.min(), x_vals.max(), 500)
            y_smooth_acumulado = spline_acumulado(x_smooth)
            y_smooth_avance_real = spline_avance_real(x_smooth)

            # Grafica las curvas suavizadas
            ax.plot(x_smooth, y_smooth_acumulado, linestyle='-', color='blue', label='Avance Programado')
            ax.plot(x_smooth, y_smooth_avance_real, linestyle='-', color='orange', label='Avance Real')

            # Grafica los puntos originales
            ax.scatter(x_vals, valores_acumulados, color='blue', zorder=5)
            ax.scatter(x_vals, avance_real_acumulado, color='orange', zorder=5)

            ax.set_title(f'Curva S Comparativa - Planilla: {planilla.name}', fontsize=16)
            ax.set_xlabel('Fecha', fontsize=12)
            ax.set_ylabel('Avance (%)', fontsize=12)
            ax.grid(True)

            yticks = ax.get_yticks()
            ax.set_yticks(yticks)
            ax.set_yticklabels([f'{tick:.0f}%' for tick in yticks], fontsize=10)

            # Anota los valores para la serie azul (avance programado) solo cuando cambian
            prev_blue = None
            for x, y in zip(x_vals, valores_acumulados):
                if prev_blue is None or y != prev_blue:
                    ax.annotate(
                        f'{y:.2f}%',
                        (x, y),
                        textcoords="offset points",
                        xytext=(-20, 15),
                        ha='center',
                        fontsize=10,
                        color='black',
                        bbox=dict(boxstyle="round,pad=0.3", fc='blue', ec='blue', alpha=0.5)
                    )
                    prev_blue = y

            # Anota los valores para la serie naranja (avance real) solo cuando cambian
            prev_orange = None
            for x, y in zip(x_vals, avance_real_acumulado):
                if prev_orange is None or y != prev_orange:
                    ax.annotate(
                        f'{y:.2f}%',
                        (x, y),
                        textcoords="offset points",
                        xytext=(20, -15),
                        ha='center',
                        fontsize=10,
                        color='black',
                        bbox=dict(boxstyle="round,pad=0.3", fc='orange', ec='orange', alpha=0.5)
                    )
                    prev_orange = y

            # Configura las etiquetas del eje X para mostrar las fechas originales
            ax.set_xticks(x_vals)
            ax.set_xticklabels([fecha.strftime('%d\n%b\n%Y') for fecha in fechas], rotation=0, fontsize=10)
            ax.legend()
            fig.tight_layout()

            # Guarda la imagen en un buffer y crea un attachment en Odoo
            img_buffer = io.BytesIO()
            fig.savefig(img_buffer, format='png', bbox_inches='tight')
            img_buffer.seek(0)

            attachment = self.env['ir.attachment'].create({
                'name': f'Curva S {planilla.name}.png',
                'type': 'binary',
                'datas': base64.b64encode(img_buffer.read()).decode('utf-8'),
                'res_model': self._name,
                'res_id': planilla.id,
                'mimetype': 'image/png',
            })

            return {
                'type': 'ir.actions.act_url',
                'url': f'/web/content/{attachment.id}?download=true',
                'target': 'new',
            }

    # Restricción SQL para asegurar que el nombre de la planilla sea único
    _sql_constraints = [
        ('unique_planilla_name', 'unique(name)', 'El nombre de la planilla debe ser único.'),
    ]

class Notebook(models.Model):
    _name = 'cronograma.notebook'
    _description = 'Notebook'

    # Relación Many2one con Planilla
    planilla_id = fields.Many2one('cronograma.planilla', string='Planilla', required=True, ondelete='cascade')
    # Campo de rubro ahora es Char (texto) y no Many2one a product.product
    rubro = fields.Char(string='Rubro', tracking=True)
    descripcion = fields.Text(string='Descripción', tracking=True)
    # Unidad de medida se selecciona manualmente (Many2one a uom.uom)
    unidad_medida = fields.Many2one('uom.uom', string='Unidad')
    # Precio unitario se ingresa manualmente
    precio_unitario = fields.Monetary(string='PVP', tracking=True)

    cantidad = fields.Float(string='Cantidad', tracking=True)
    # Campos calculados basados en precio unitario y cantidad/consumo
    subtotal = fields.Monetary(string='Sub. Valorado', compute='_compute_subtotal', store=True, readonly=True, tracking=True)
    diferencia_subtotal = fields.Monetary(string='Diferencia subtotal', compute='_compute_diferencia_subtotal', store=True, readonly=True, tracking=True)
    consumo = fields.Float(string='Consumo', compute='_compute_consumo', store=True, readonly=True)
    diferencia_cantidad = fields.Float(string='Diferencia consumo', compute='_compute_diferencia_cantidad', store=True, readonly=True, tracking=True)
    subtotal_planilla = fields.Monetary(string='Sub. Avance Real', compute='_compute_subtotal_planilla', store=True, readonly=True, tracking=True)

    # Relación uno a muchos con los detalles del notebook
    detalle_ids = fields.One2many('cronograma.detalle', 'notebook_id', string='Detalles', tracking=True)
    observacion = fields.Text(string='Observación', tracking=True)

    # Moneda derivada de la planilla
    currency_id = fields.Many2one(related="planilla_id.currency_id", string="Moneda", store=False, readonly=True)
    nivel = fields.Integer(string='Nivel', readonly=True)
    
    # Se eliminó el metodo onchange, ya que ahora se ingresa el precio manualmente

    @api.depends('detalle_ids.valor')
    def _compute_consumo(self):
        """Calcula el consumo sumando la cantidad de avance real de cada detalle."""
        for record in self:
            record.consumo = sum(detalle.cantidad_avance_real for detalle in record.detalle_ids)

    @api.depends('precio_unitario', 'cantidad')
    def _compute_subtotal(self):
        """Calcula el subtotal valorado multiplicando el precio unitario por la cantidad."""
        for record in self:
            record.subtotal = record.precio_unitario * record.cantidad if record.cantidad > 0 else 0

    @api.depends('subtotal', 'subtotal_planilla')
    def _compute_diferencia_subtotal(self):
        """Calcula la diferencia entre el subtotal valorado y el subtotal de avance real."""
        for record in self:
            record.diferencia_subtotal = record.subtotal - record.subtotal_planilla

    @api.depends('precio_unitario', 'consumo')
    def _compute_subtotal_planilla(self):
        """Calcula el subtotal de avance real multiplicando el precio unitario por el consumo."""
        for record in self:
            record.subtotal_planilla = record.precio_unitario * record.consumo if record.consumo > 0 else 0

    @api.depends('cantidad', 'consumo')
    def _compute_diferencia_cantidad(self):
        """Calcula la diferencia entre la cantidad total y el consumo realizado."""
        for record in self:
            record.diferencia_cantidad = record.cantidad - record.consumo

class Detalle(models.Model):
    _name = 'cronograma.detalle'
    _description = 'Detalle del Notebook'

    # Relación Many2one con Notebook
    notebook_id = fields.Many2one('cronograma.notebook', string='Notebook', required=True, ondelete='cascade')
    fecha = fields.Date(string='Fecha', required=True, readonly=True)  # Fecha del detalle
    cantidad = fields.Float(string='Cantidad', required=True)  # Cantidad programada para el detalle
    cantidad_avance_real = fields.Float(string='Cantidad Avance Real', required=True)  # Cantidad efectivamente realizada

    # Moneda derivada del notebook
    currency_id = fields.Many2one(related='notebook_id.currency_id', string="Moneda", store=False, readonly=True)

    # Campos calculados para el valor, y el valor de avance real
    valor = fields.Monetary(string='Valor', compute='_compute_valor', store=True, readonly=True)
    valor_avance_real = fields.Monetary(string='Valor Avance Real', compute='_compute_valor_avance_real', store=True, readonly=True)

    @api.depends('cantidad', 'notebook_id.precio_unitario')
    def _compute_valor(self):
        """Calcula el valor del detalle multiplicando la cantidad por el precio unitario."""
        for record in self:
            record.valor = record.cantidad * record.notebook_id.precio_unitario

    @api.depends('cantidad_avance_real', 'notebook_id.precio_unitario')
    def _compute_valor_avance_real(self):
        """Calcula el valor de avance real multiplicando la cantidad de avance real por el precio unitario."""
        for record in self:
            record.valor_avance_real = record.cantidad_avance_real * record.notebook_id.precio_unitario
